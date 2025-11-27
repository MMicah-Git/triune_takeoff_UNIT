# triune_takeoff_manual_only.py
# Triune Takeoff — Manual product & per-product tag ordering (textarea only)
# Run:
#   python -m pip install streamlit pandas openpyxl
#   streamlit run triune_takeoff_manual_only.py

import io
import re
import json
import hashlib
import zipfile
from typing import Dict, List, Optional, Tuple

import pandas as pd
import numpy as np
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------
# App constants
# ---------------------------
APP_TITLE = "Triune Takeoff — Manual ordering (textarea only)"

TRIUNE_COLUMNS = [
    "PRODUCT", "BRAND", "MODEL", "QTY", "TAG",
    "NECK SIZE", "MODULE SIZE", "DUCT SIZE", "TYPE", "MOUNTING",
    "ACCESSORIES1", "ACCESSORIES2", "REMARK"
]

DEFAULT_COLORS = {
    "header": "#ECF3FA",
    "model":  "#F4B084",
    "product": "#92D050",
    "zebra":  "#F7F7F7",
    "qtygold": "#FFF2CC"
}

BOLD_COLUMNS_DEFAULT = "PRODUCT,BRAND,TAG,MODULE SIZE,TYPE,ACCESSORIES1"

STRICT_KEYS = [
    "PRODUCT", "BRAND", "MODEL", "TAG",
    "NECK SIZE", "MODULE SIZE", "DUCT SIZE", "TYPE", "MOUNTING"
]

CUSTOMER_ABBR = {
    "Gustave Larson": "GAL",
    "Southvac Systems": "SS",
    "Midwest Mechanical Solutions": "MMS",
    "Haldeman": "HM",
    "Knape Associates": "KA",
    "Knape Dallas": "KAD",
    "Knape Houston": "KAH",
    "Applied Product Solutions": "APS",
}

# ---------------------------
# Utilities
# ---------------------------
def strip(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    return str(x).strip()

def sanename(s: str) -> str:
    s = str(s).strip().replace(" ", "_")
    return re.sub(r"[^A-Za-z0-9._-]+", "", s)

def normalize_product_text(s: str) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    return t.lower()

def _norm_key(s: str) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()

def read_uploaded_file(uploaded_file):
    if uploaded_file is None:
        return None, "No file uploaded."
    name = uploaded_file.name.lower()
    if name.endswith(('.xls', '.xlsx')):
        try:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, dtype=str)
            return df.fillna(''), None
        except Exception as e:
            return None, f"Failed to read Excel file: {e}"
    try:
        uploaded_file.seek(0)
        raw = uploaded_file.read()
        if isinstance(raw, bytes):
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                try:
                    text = raw.decode('latin1')
                except Exception:
                    return None, "Failed to decode CSV bytes as utf-8 or latin1."
        else:
            text = str(raw)
        lines = [ln for ln in text.splitlines() if ln.strip() != '']
        if not lines:
            return None, "File appears empty."
        sample = "\n".join(lines[:10])
        try:
            import csv as _csv
            dialect = _csv.Sniffer().sniff(sample)
            delim = dialect.delimiter
        except Exception:
            for d in [',', ';', '\t', '|']:
                if d in sample:
                    delim = d
                    break
            else:
                delim = ','
        try:
            df = pd.read_csv(io.StringIO(text), sep=delim, dtype=str, engine='python')
        except Exception:
            df = pd.read_csv(io.StringIO(text), sep=None, engine='python', dtype=str)
        return df.fillna(''), None
    except Exception as e:
        return None, f"Unexpected error reading file: {e}"

# ---------------------------
# Tag parsing / sort helper
# ---------------------------
def parse_tag_for_sort(tag: str) -> Tuple[int, Optional[int], Optional[str], str]:
    t = str(tag).strip()
    if t == "":
        return (3, None, None, t)
    if re.fullmatch(r"\d+", t):
        try:
            return (0, int(t.lstrip("0") or "0"), None, t)
        except Exception:
            return (3, None, None, t)
    if re.fullmatch(r"[A-Za-z]+", t):
        return (1, None, t.upper(), t)
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", t)
    if m:
        letters = m.group(1).upper()
        num = int(m.group(2).lstrip("0") or "0")
        return (2, num, letters, t)
    return (3, None, None, t)

# ---------------------------
# Triune helper functions (TAG padding fixed to 2 digits)
# ---------------------------
def detect_map(df: pd.DataFrame) -> Dict[str, str]:
    cols = {c.lower(): c for c in df.columns}
    def find(*names):
        for n in names:
            if n.lower() in cols:
                return cols[n.lower()]
        for key, orig in cols.items():
            for n in names:
                if n.lower() in key:
                    return orig
        return ""
    return {
        "PRODUCT": find("subject", "product", "item", "product name", "prod"),
        "BRAND": find("manufacturer", "brand", "make", "mfr"),
        "MODEL": find("model", "catalog", "cat no", "catalog no"),
        "QTY": find("quantity", "qty", "count", "qty.", "q'ty"),
        "TAG": find("label", "tag", "ref", "mark"),
        "NECK SIZE": find("neck size", "neck"),
        "MODULE SIZE": find("module size", "face size", "module", "face"),
        "DUCT SIZE": find("duct size", "duct"),
        "TYPE": find("type", "desc", "description"),
        "MOUNTING": find("mounting", "install"),
        "ACCESSORIES1": find("accessories", "accessories1", "accessory 1"),
        "ACCESSORIES2": find("accessories2", "accessory 2", "description"),
        "REMARK": find("remark", "remarks", "note", "notes")
    }

def neck_num(v):
    m = re.search(r"\d+(?:\.\d+)?", str(v))
    return float(m.group()) if m else np.inf

def _numbers_from_text(s: str):
    if s is None:
        return []
    s = str(s)
    s = re.sub(r"[×xX]", "x", s)
    s = re.sub(r"(?<=\d),(?=\d)", "", s)
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    return [float(n) for n in nums]

def size_pair(s: str):
    nums = _numbers_from_text(s)
    if not nums:
        return (float("inf"), float("inf"))
    if len(nums) == 1:
        return (nums[0], 0.0)
    return (nums[0], nums[1])

def normalize(df: pd.DataFrame, mapping: Dict[str, str]):
    out = pd.DataFrame()
    for c in TRIUNE_COLUMNS:
        src = mapping.get(c, "")
        out[c] = df[src] if src in df.columns else ""
    out["QTY"] = pd.to_numeric(out["QTY"], errors="coerce").fillna(0)
    for c in out.columns:
        if c != "QTY":
            out[c] = out[c].map(strip)

    out["_NECK_NUM"] = out["NECK SIZE"].map(neck_num)
    mod = out["MODULE SIZE"].map(size_pair)
    dct = out["DUCT SIZE"].map(size_pair)
    out["_MODULE_W"] = [t[0] for t in mod]
    out["_MODULE_H"] = [t[1] for t in mod]
    out["_DUCT_W"]   = [t[0] for t in dct]
    out["_DUCT_H"]   = [t[1] for t in dct]

    # pad numeric TAGs to 2 digits
    tag_series = out["TAG"].astype(str).fillna("").map(str).map(lambda s: s.strip())
    def pad_2(tv):
        tvs = str(tv).strip()
        if re.fullmatch(r"\d+", tvs):
            return tvs.zfill(2)
        return tvs
    out["TAG"] = tag_series.map(pad_2)

    return out, mapping

def apply_grouping_strict(df: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        df.groupby(STRICT_KEYS, dropna=False)
          .agg({
              "QTY":"sum",
              "_NECK_NUM":"min","_MODULE_W":"min","_MODULE_H":"min","_DUCT_W":"min","_DUCT_H":"min",
              "ACCESSORIES1":"first","ACCESSORIES2":"first","REMARK":"first"
          }).reset_index()
    )
    grouped = grouped.sort_values(
        by=["PRODUCT","MODEL","TAG","_NECK_NUM","_MODULE_W","_MODULE_H","_DUCT_W","_DUCT_H","TYPE","MOUNTING","BRAND"]
    ).reset_index(drop=True)
    return grouped

def build_rows_toprow(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for p, Gp in df.groupby("PRODUCT", dropna=False):
        for (prod, brand, model), g in Gp.groupby(["PRODUCT","BRAND","MODEL"], dropna=False):
            first=True
            for _, r in g.iterrows():
                rec = {c: r.get(c,"") for c in TRIUNE_COLUMNS}
                if not first:
                    rec.update({"PRODUCT":"", "BRAND":"", "MODEL":""})
                first=False
                rows.append(rec)
            rows.append({**{c:"" for c in TRIUNE_COLUMNS}, "MODEL": f"{model} Total", "QTY": float(g["QTY"].sum())})
        rows.append({**{c:"" for c in TRIUNE_COLUMNS}, "PRODUCT": f"{p} Total", "QTY": float(Gp["QTY"].sum())})
    rows.append({**{c:"" for c in TRIUNE_COLUMNS}, "PRODUCT": "Grand Total", "QTY": float(df["QTY"].sum())})
    return pd.DataFrame(rows)[TRIUNE_COLUMNS]

def style_to_excel_bytes(df: pd.DataFrame,
                         header_hex, model_hex, product_hex, zebra_hex, qty_gold_hex,
                         bold_cols: List[str]) -> bytes:
    def hex_to_fg(h): return h.replace("#", "").upper()

    wb = Workbook()
    ws = wb.active
    ws.title = "Triune Output"
    ws.freeze_panes = "A2"

    ws.append(TRIUNE_COLUMNS)
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=hex_to_fg(header_hex))
    model_fill  = PatternFill("solid", fgColor=hex_to_fg(model_hex))
    product_fill= PatternFill("solid", fgColor=hex_to_fg(product_hex))
    zebra_fill  = PatternFill("solid", fgColor=hex_to_fg(zebra_hex))
    qty_fill    = PatternFill("solid", fgColor=hex_to_fg(qty_gold_hex))

    qty_idx = TRIUNE_COLUMNS.index("QTY")+1
    prod_idx= 1
    brand_idx=2
    model_idx=3

    for c in range(1, len(TRIUNE_COLUMNS)+1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="000000")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r in range(2, ws.max_row+1):
        v_prod = (ws.cell(r, prod_idx).value or "")
        v_mod  = (ws.cell(r, model_idx).value or "")
        is_pt = isinstance(v_prod, str) and v_prod.endswith(" Total") and v_prod != "Grand Total"
        is_gt = v_prod == "Grand Total"
        is_mt = isinstance(v_mod, str) and v_mod.endswith(" Total")

        for c in range(1, len(TRIUNE_COLUMNS)+1):
            col_name = TRIUNE_COLUMNS[c-1]
            cell = ws.cell(r, c)
            cell.border = border
            cell.font = Font(bold=(col_name in set(bold_cols)), color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if c == qty_idx:
                cell.number_format = '#,##0'

        if is_mt:
            for c in range(1, len(TRIUNE_COLUMNS)+1):
                if c in (prod_idx, brand_idx): continue
                ws.cell(r, c).fill = model_fill
                ws.cell(r, c).font = Font(bold=True, color="000000")
        elif is_pt:
            for c in range(1, len(TRIUNE_COLUMNS)+1):
                ws.cell(r, c).fill = product_fill
                ws.cell(r, c).font = Font(bold=True, color="000000")
        elif is_gt:
            for c in range(1, len(TRIUNE_COLUMNS)+1):
                ws.cell(r, c).fill = header_fill
                ws.cell(r, c).font = Font(bold=True, color="000000")
        else:
            if r % 2 == 0:
                for c in range(1, len(TRIUNE_COLUMNS)+1):
                    ws.cell(r, c).fill = zebra_fill
            ws.cell(r, qty_idx).fill = qty_fill

    for i in range(1, len(TRIUNE_COLUMNS)+1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def bytes_from_df_excel(df: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data Unit Sheet")
    return out.getvalue()

def create_zip_bytes_from_map(dfs_map: dict, export_fn) -> bytes:
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for unit, df in dfs_map.items():
            safe = "".join(c if c.isalnum() or c in " -_." else "_" for c in str(unit))[:120] or "unit"
            fname = f"{safe}.xlsx"
            zf.writestr(fname, export_fn(df))
    return mem.getvalue()

def export_styled_excel_bytes(df: pd.DataFrame) -> bytes:
    return bytes_from_df_excel(df)

# ---------------------------
# Helpers to safely rerun (compatibility)
# ---------------------------
def safe_rerun():
    """
    Attempt to trigger a rerun in Streamlit. Some Streamlit builds may not expose
    st.experimental_rerun; this wrapper swallows the AttributeError (no-op).
    """
    try:
        rerun_fn = getattr(st, "experimental_rerun", None)
        if callable(rerun_fn):
            rerun_fn()
            return
        try:
            from streamlit.runtime.scriptrunner import RerunException
            raise RerunException()
        except Exception:
            return
    except Exception:
        return

# ---------------------------
# Manual-only Editors with Apply buttons
# ---------------------------
def render_manual_product_editor(preview_df: pd.DataFrame):
    """
    Manual product order editor: textarea only (one product per line).
    """
    product_values = preview_df["PRODUCT"].fillna("").astype(str).tolist()
    seen = set()
    products = []
    for p in product_values:
        if p not in seen:
            seen.add(p)
            products.append(p)

    if not products:
        st.info("No products found to order.")
        return

    with st.expander("Manual PRODUCT ordering — one product per line (click Apply to save)", expanded=True):
        fallback_key = "manual_order_textarea"
        default_text = "\n".join(products)
        # prefill with existing saved original order if present, else detected products
        pref = st.session_state.get('triune_manual_products_original')
        if pref and isinstance(pref, list) and len(pref) > 0:
            default_text = "\n".join(pref)
        pasted = st.text_area("Manual product order (one product per line):", value=st.session_state.get(fallback_key, default_text), height=240, key=fallback_key)
        # do not assign to st.session_state[fallback_key] after creating the widget
        lines = [ln.strip() for ln in pasted.splitlines() if ln.strip() != ""]
        preview_order = lines or products
        st.markdown("**Preview (unsaved) product order:**")
        st.write(preview_order)

        # Build normalized manual list but do not save until Apply clicked
        manual_norm = []
        manual_orig = []
        seen2 = set()
        for orig in preview_order:
            pnorm = normalize_product_text(orig)
            if pnorm != "" and pnorm not in seen2:
                seen2.add(pnorm)
                manual_norm.append(pnorm)
                manual_orig.append(orig)

        col1, col2 = st.columns([1,1])
        if col1.button("Apply product order"):
            if manual_norm:
                st.session_state['triune_manual_products'] = manual_norm
                st.session_state['triune_manual_products_original'] = manual_orig
                st.success(f"Product order applied — {len(manual_norm)} products saved.")
                safe_rerun()
            else:
                st.session_state.pop('triune_manual_products', None)
                st.session_state.pop('triune_manual_products_original', None)
                st.info("Cleared manual product order (nothing saved).")

        if col2.button("Reset manual product order"):
            st.session_state.pop('triune_manual_products', None)
            st.session_state.pop('triune_manual_products_original', None)
            st.success("Manual product order cleared.")
            safe_rerun()

def render_manual_per_product_tag_editor(preview_df: pd.DataFrame):
    """
    Per-product TAG editor: textarea only (one tag per line).
    """
    detected_products = preview_df["PRODUCT"].fillna("").astype(str).tolist()
    seen = set(); prod_list = []
    for p in detected_products:
        pn = _norm_key(p)
        if pn not in seen:
            seen.add(pn); prod_list.append(p)

    manual_prod_norm = st.session_state.get('triune_manual_products') or []
    manual_prod_display = []
    if manual_prod_norm:
        orig_map = {}
        for orig in prod_list:
            orig_map[_norm_key(orig)] = orig
        for mp in manual_prod_norm:
            if mp in orig_map:
                manual_prod_display.append(orig_map[mp])
        for p in prod_list:
            if p not in manual_prod_display:
                manual_prod_display.append(p)
        products_for_ui = manual_prod_display
    else:
        products_for_ui = prod_list

    if not products_for_ui:
        st.info("No products found for per-product tag editor.")
        return

    prev = st.session_state.get('triune_perprod_selected_product')
    sel_index = 0
    if prev and prev in products_for_ui:
        sel_index = products_for_ui.index(prev)
    sel_prod = st.selectbox("Select product to edit tags for", options=products_for_ui, index=sel_index, key="perprod_tag_select")
    st.session_state['triune_perprod_selected_product'] = sel_prod
    sel_prod_norm = _norm_key(sel_prod)

    # collect tags for selected product
    tags_raw = preview_df.loc[preview_df["PRODUCT"].fillna("").astype(str) == sel_prod, "TAG"].fillna("").astype(str).tolist()
    seen_t = set(); tags = []
    for t in tags_raw:
        if t not in seen_t:
            seen_t.add(t); tags.append(t)

    if not tags:
        st.info(f"No TAGs found for product: {sel_prod}")
        return

    with st.expander(f"Edit TAG order for product: {sel_prod} (click Apply to save)", expanded=True):
        fallback_key = f"manual_tag_order_{_norm_key(sel_prod)}"
        default_text = "\n".join(tags)
        # if an original saved list exists, use that as default
        saved_orig_map = st.session_state.get('triune_manual_tags_by_product_orig') or {}
        if sel_prod_norm in saved_orig_map:
            default_text = "\n".join(saved_orig_map[sel_prod_norm])
        pasted = st.text_area("Manual tag order for this product (one tag per line):", value=st.session_state.get(fallback_key, default_text), height=200, key=fallback_key)
        lines = [ln.strip() for ln in pasted.splitlines() if ln.strip() != ""]
        preview_order = lines or tags
        st.markdown("**Preview (unsaved) tag order for product:**")
        st.write(preview_order)

        # build normalized lists but do not save until Apply clicked
        manual_norm_list = []
        manual_orig_list = []
        for tv in preview_order:
            tn = _norm_key(tv)
            if tn != "":
                manual_norm_list.append(tn)
                manual_orig_list.append(tv)

        col1, col2 = st.columns([1,1])
        if col1.button(f"Apply TAG order for product: {sel_prod}"):
            if 'triune_manual_tags_by_product' not in st.session_state:
                st.session_state['triune_manual_tags_by_product'] = {}
            if 'triune_manual_tags_by_product_orig' not in st.session_state:
                st.session_state['triune_manual_tags_by_product_orig'] = {}
            if manual_norm_list:
                st.session_state['triune_manual_tags_by_product'][sel_prod_norm] = manual_norm_list
                st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm] = manual_orig_list
                st.success(f"Saved manual TAG order for product: {sel_prod} ({len(manual_norm_list)} tags).")
                safe_rerun()
            else:
                if sel_prod_norm in st.session_state.get('triune_manual_tags_by_product', {}):
                    del st.session_state['triune_manual_tags_by_product'][sel_prod_norm]
                if sel_prod_norm in st.session_state.get('triune_manual_tags_by_product_orig', {}):
                    del st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm]
                st.info(f"Cleared manual TAG order for product: {sel_prod} (nothing saved).")
        if col2.button(f"Reset TAG order for product: {sel_prod}"):
            if 'triune_manual_tags_by_product' in st.session_state and sel_prod_norm in st.session_state['triune_manual_tags_by_product']:
                del st.session_state['triune_manual_tags_by_product'][sel_prod_norm]
            if 'triune_manual_tags_by_product_orig' in st.session_state and sel_prod_norm in st.session_state['triune_manual_tags_by_product_orig']:
                del st.session_state['triune_manual_tags_by_product_orig'][sel_prod_norm]
            st.success(f"Cleared saved TAG order for product: {sel_prod}.")
            safe_rerun()

# ---------------------------
# Ordering helpers & pipeline (unchanged logic)
# ---------------------------
def reorder_grouped(grouped: pd.DataFrame,
                    sort_keys: List[str],
                    ascending: bool = True,
                    manual_products: Optional[List[str]] = None,
                    manual_tags: Optional[object] = None) -> pd.DataFrame:
    df = grouped.copy().reset_index(drop=True)
    prod_series = df['PRODUCT'].fillna("").astype(str).tolist()
    tag_series  = df['TAG'].fillna("").astype(str).tolist()
    prod_norms = [_norm_key(x) for x in prod_series]
    tag_norms  = [_norm_key(x) for x in tag_series]

    # manual product ordering
    if manual_products:
        seen = set()
        manual_norm = []
        for p in manual_products:
            p_norm = _norm_key(p)
            if p_norm and p_norm not in seen:
                seen.add(p_norm); manual_norm.append(p_norm)
        unique_prods = []
        unique_prods_norm = []
        for orig, norm in zip(prod_series, prod_norms):
            if norm not in unique_prods_norm:
                unique_prods_norm.append(norm); unique_prods.append(orig)
        manual_in_df = []
        for m in manual_norm:
            if m in unique_prods_norm:
                manual_in_df.append(unique_prods[unique_prods_norm.index(m)])
        if manual_in_df:
            remaining = [p for p in unique_prods if p not in manual_in_df]
            cat_order = manual_in_df + remaining
            df['PRODUCT'] = pd.Categorical(df['PRODUCT'], categories=cat_order, ordered=True)
            if sort_keys is None or (len(sort_keys)==0) or (sort_keys[0] != 'PRODUCT'):
                sort_keys = ['PRODUCT'] + [k for k in (sort_keys or []) if k != 'PRODUCT']

    # manual_tags expected to be per-product dict (normalized keys)
    manual_by_product = {}
    if manual_tags and isinstance(manual_tags, dict):
        for k,v in manual_tags.items():
            manual_by_product[_norm_key(k)] = [_norm_key(x) for x in (v or [])]

    # auto order baseline
    auto_orders = []
    for t in tag_series:
        cat, num, alpha, orig = parse_tag_for_sort(t)
        if cat == 0:
            auto_orders.append( (0, num if num is not None else 0, "", "") )
        elif cat == 1:
            auto_orders.append( (1, 0, alpha or "", "") )
        elif cat == 2:
            auto_orders.append( (2, num if num is not None else 0, alpha or "", "") )
        else:
            auto_orders.append( (3, 0, "", t) )
    order_idx = sorted(range(len(auto_orders)), key=lambda i: auto_orders[i])
    auto_rank = {i: rank for rank, i in enumerate(order_idx)}

    pri_list = []
    for i, (p_norm, t_norm) in enumerate(zip(prod_norms, tag_norms)):
        L = manual_by_product.get(p_norm)
        if L:
            if t_norm in L:
                pri = L.index(t_norm)
            else:
                pri = 1000000 + auto_rank.get(i, 0)
        else:
            pri = 1000000 + auto_rank.get(i, 0)
        pri_list.append(pri)

    df['_TAG_PRI'] = pri_list

    sk = sort_keys or []
    effective_sort_keys = []
    for k in sk:
        if k == 'TAG':
            effective_sort_keys.append('_TAG_PRI')
        else:
            effective_sort_keys.append(k)

    if not effective_sort_keys:
        effective_sort_keys = ["PRODUCT", "MODEL", "_TAG_PRI", "_NECK_NUM", "_MODULE_W", "_MODULE_H", "_DUCT_W", "_DUCT_H"]

    valid_keys = [k for k in effective_sort_keys if k in df.columns]
    if len(valid_keys) == 0:
        df = df.drop(columns=[c for c in ['_TAG_PRI'] if c in df.columns])
        return df

    df = df.sort_values(by=valid_keys, ascending=ascending, kind='mergesort').reset_index(drop=True)
    if '_TAG_PRI' in df.columns:
        df = df.drop(columns=['_TAG_PRI'])
    return df

def takeoff_pipeline(df_raw: pd.DataFrame,
                     mapping: Dict[str,str],
                     sort_keys: Optional[List[str]] = None,
                     ascending: bool = True,
                     manual_products: Optional[List[str]] = None,
                     manual_tags: Optional[object] = None) -> pd.DataFrame:
    norm, _ = normalize(df_raw, mapping)
    grouped = apply_grouping_strict(norm)
    grouped_reordered = reorder_grouped(grouped,
                                        sort_keys or [],
                                        ascending=ascending,
                                        manual_products=manual_products,
                                        manual_tags=manual_tags)
    rows = build_rows_toprow(grouped_reordered)
    cols = [c for c in rows.columns if not c.startswith("_")]
    return rows[cols]

# ---------------------------
# Data Unit helpers (unchanged)
# ---------------------------
def normalize_text(s: str) -> str:
    return re.sub(r"[\s_\-\.]+", " ", str(s).strip().lower())

def find_column(raw_cols: List[str], *aliases: str) -> Optional[str]:
    raw_norm = {c: normalize_text(c) for c in raw_cols}
    for alias in aliases:
        a_norm = normalize_text(alias)
        for rc, rn in raw_norm.items():
            if rn == a_norm:
                return rc
    for alias in aliases:
        a_norm = normalize_text(alias)
        for rc, rn in raw_norm.items():
            if a_norm in rn or rn in a_norm:
                return rc
    alias_tokens = set()
    for alias in aliases:
        alias_tokens.update(normalize_text(alias).split())
    best = None
    best_score = 0
    for rc, rn in raw_norm.items():
        tokens = set(rn.split())
        score = len(tokens & alias_tokens)
        if score > best_score:
            best_score = score
            best = rc
    if best_score > 0:
        return best
    return None

def load_file_to_df_simple(uploaded_file):
    if uploaded_file is None:
        return None
    name = getattr(uploaded_file, "name", str(uploaded_file)).lower()
    try:
        if name.endswith(".csv"):
            return pd.read_csv(uploaded_file, dtype=str)
        else:
            return pd.read_excel(uploaded_file, sheet_name=0, dtype=str)
    except Exception as e:
        st.error(f"Error reading file {name}: {e}")
        return None

def detect_unit_column(df: pd.DataFrame) -> Optional[str]:
    for c in df.columns:
        if re.search(r"(^unit$)|unit(s)?|apt|apartment|flat|room|suite|unit_id|unitno|unit_no", c, re.IGNORECASE):
            return c
    return None

def apply_raw_column_mapping(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = raw_df.copy()
    cols = list(df.columns)
    lowered = {c.lower(): c for c in cols}
    exact_rules = {
        "subject": "PRODUCT",
        "page index": "Page Index",
        "label": "TAG",
        "manufacturer": "BRAND",
        "face size": "MODULE SIZE",
        "description": "ACCESSORIES2",
        "accessories": "ACCESSORIES1",
        "accessories1": "ACCESSORIES1"
    }
    exact_map = {}
    for low_key, dest in exact_rules.items():
        if low_key in lowered:
            orig = lowered[low_key]
            if orig != dest:
                exact_map[orig] = dest
    if exact_map:
        df = df.rename(columns=exact_map)
        cols = list(df.columns)
    mapping_pairs = {
        "QTY": ("quantity", "qty", "count", "qty.", "q'ty"),
        "BRAND": ("brand", "make", "mfr"),
        "MODEL": ("model", "catalog", "cat no", "catalog no"),
        "TAG": ("tag", "tag id", "label", "ref", "mark"),
        "NECK SIZE": ("neck size", "neck"),
        "MODULE SIZE": ("module size", "face size", "module", "face"),
        "DUCT SIZE": ("duct size", "duct"),
        "CFM": ("cfm", "airflow"),
        "TYPE": ("type", "desc", "description"),
        "MOUNTING": ("mounting", "install", "mount"),
        "ACCESSORIES1": ("accessories", "accessories1", "accessory 1", "accessory"),
        "ACCESSORIES2": ("accessories2", "accessory 2", "description", "desc"),
        "REMARK": ("remark", "remarks", "note", "notes"),
        "UNITS": ("unit", "units", "zone", "area", "apt", "apartment")
    }
    rename_dict = {}
    cols = list(df.columns)
    for target_col, aliases in mapping_pairs.items():
        if target_col in df.columns:
            continue
        found = find_column(cols, *aliases)
        if found and found not in rename_dict and found != target_col:
            rename_dict[found] = target_col
    if rename_dict:
        df = df.rename(columns=rename_dict)
    return df

def clean_unit_matrix(df: pd.DataFrame, unit_col_hint: Optional[str] = None) -> (pd.DataFrame, str):
    df2 = df.copy()
    unit_col = unit_col_hint or detect_unit_column(df2) or df2.columns[0]
    df2[unit_col] = df2[unit_col].fillna("").astype(str).str.strip()
    df2 = df2[~df2[unit_col].str.upper().isin(["TOTAL", "GRAND TOTAL", "SUMMARY", "ALL"])]
    return df2.reset_index(drop=True), unit_col

def guess_multiplier_column(unit_df: pd.DataFrame, unit_col: str) -> Optional[str]:
    candidates = [c for c in unit_df.columns if c != unit_col]
    for c in candidates:
        sample = unit_df[c].dropna().astype(str).str.replace(",", "").str.strip()
        if sample.size and any(s.replace(".", "", 1).isdigit() for s in sample[:50]):
            return c
    return candidates[0] if candidates else None

def build_data_unit_sheet(
    raw_df: pd.DataFrame,
    unit_df: pd.DataFrame,
    raw_unit_col: str,
    matrix_unit_col: str,
    multiplier_col: Optional[str],
    selected_units: Optional[list] = None,
    include_empty: bool = True,
    default_multiplier: int = 1
) -> pd.DataFrame:
    raw = raw_df.copy()
    mat = unit_df.copy()
    raw[raw_unit_col] = raw[raw_unit_col].fillna("").astype(str).str.strip()
    mat[matrix_unit_col] = mat[matrix_unit_col].fillna("").astype(str).str.strip()
    multiplier_map = {}
    if multiplier_col:
        for _, r in mat.iterrows():
            u = str(r.get(matrix_unit_col, "")).strip()
            v = r.get(multiplier_col, "")
            try:
                if pd.isna(v) or str(v).strip() == "":
                    continue
                num = int(float(str(v).replace(",", "").strip()))
                multiplier_map[u] = num
            except Exception:
                continue
    org_count_col = None
    for name in ["Org Count", "OrgCount", "ORIGINAL COUNT", "Count", "QTY", "Qty", "qty", "QTY/UNIT"]:
        if name in raw.columns:
            org_count_col = name
            break
    if org_count_col is None:
        for c in raw.columns:
            sample = raw[c].dropna().astype(str).str.replace(",", "").str.strip()
            if sample.size and all(s.replace(".", "", 1).isdigit() for s in sample[:50]):
                org_count_col = c
                break
    if org_count_col is None:
        raw["Org Count"] = 1
    else:
        raw["Org Count"] = pd.to_numeric(raw[org_count_col].fillna("0").astype(str).str.replace(",", "").str.strip(), errors="coerce").fillna(0).astype(int)
    def get_mult(u):
        if u == "" or pd.isna(u):
            return default_multiplier
        return multiplier_map.get(u, default_multiplier)
    raw["__unit_multiplier__"] = raw[raw_unit_col].apply(get_mult).astype(int)
    raw["Count"] = raw["Org Count"].astype(int) * raw["__unit_multiplier__"]
    if selected_units:
        sel = set(selected_units)
        raw_units_vals = raw[raw_unit_col].fillna("").astype(str).str.strip()
        mask = raw_units_vals.isin(sel)
        if "<<EMPTY UNIT>>" in sel:
            mask = mask | (raw_units_vals == "")
        raw = raw[mask].copy()
    out_unit_name = "UNITS"
    raw[out_unit_name] = raw[raw_unit_col].replace({"": "<<EMPTY UNIT>>"})
    required_order = [
        "PRODUCT", "Page Index", "TAG", "Org Count", "Count", "BRAND", "MODEL",
        "NECK SIZE", "MODULE SIZE", "DUCT SIZE", "CFM", "TYPE", "MOUNTING",
        "ACCESSORIES1", "ACCESSORIES2", "REMARK", out_unit_name, "DAMPER TYPE"
    ]
    out = raw.copy()
    for col in required_order:
        if col not in out.columns:
            out[col] = ""
    other_cols = [c for c in out.columns if c not in required_order and c != "__unit_multiplier__"]
    final_cols = required_order + other_cols
    if "__unit_multiplier__" in out.columns:
        out = out.drop(columns=["__unit_multiplier__"])
    return out[final_cols].copy()

# ---------------------------
# Streamlit UI
# ---------------------------
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)

# Sidebar (hidden controls)
st.sidebar.header("Triune Options (hidden)")
header_color = st.sidebar.color_picker("Header / Grand Total", DEFAULT_COLORS["header"], key="hdrcol")
model_color = st.sidebar.color_picker("Model Total", DEFAULT_COLORS["model"], key="modcol")
product_color = st.sidebar.color_picker("Product Total", DEFAULT_COLORS["product"], key="prodcol")
zebra_color = st.sidebar.color_picker("Alternating row", DEFAULT_COLORS["zebra"], key="zebcol")
qty_color = st.sidebar.color_picker("QTY (detail rows)", DEFAULT_COLORS["qtygold"], key="qtycol")

st.sidebar.markdown("---")
st.sidebar.subheader("Mapping JSON (optional)")
mapping_json = st.sidebar.text_area("Mapping JSON (target -> source). Leave empty to auto-detect.", value="", height=160, key="mapjson")
if st.sidebar.button("Load mapping from JSON", key="loadmapbtn"):
    try:
        m = json.loads(mapping_json)
        st.session_state.triune_mapping = m
        st.sidebar.success("Mapping loaded and persisted.")
    except Exception as e:
        st.sidebar.error(f"Invalid JSON: {e}")

tabs = st.tabs(["Triune Takeoff", "Data Unit Sheet"])

# ========== TAB 0: Triune Takeoff ==========
with tabs[0]:
    st.header("Triune Takeoff — Convert raw takeoff to Triune format")
    uploaded = st.file_uploader("Upload CSV or Excel (raw takeoff)", type=['csv','xlsx','xls'], key="triune_upload")

    st.markdown("---")
    st.subheader("Options (main)")
    bold_input = st.text_input("Columns to bold", value=BOLD_COLUMNS_DEFAULT, key="boldcols")
    bold_cols = [c.strip() for c in bold_input.split(",") if c.strip()]

    if uploaded:
        df_in, read_err = read_uploaded_file(uploaded)
        if read_err:
            st.error(f"Failed to read file: {read_err}")
        else:
            if "triune_mapping" not in st.session_state:
                st.session_state.triune_mapping = None
            if "upload_sig_triune" not in st.session_state:
                st.session_state.upload_sig_triune = None

            def file_signature(df, name):
                cols = list(df.columns) if df is not None else []
                payload = name + "|" + "|".join(cols)
                return hashlib.md5(payload.encode("utf-8")).hexdigest()

            sig = file_signature(df_in, uploaded.name)
            if st.session_state.upload_sig_triune != sig:
                st.session_state.upload_sig_triune = sig
                st.session_state.ran_triune = False
                st.session_state.formatted_df_triune = None

            auto_map = detect_map(df_in)
            mapping_base = auto_map.copy()
            if st.session_state.triune_mapping:
                for k,v in st.session_state.triune_mapping.items():
                    mapping_base[k] = v

            st.subheader("Column mapping (target → source)")
            with st.form("mapping_form_triune"):
                user_map = {}
                for c in TRIUNE_COLUMNS:
                    choices = [''] + list(df_in.columns)
                    default = mapping_base.get(c) if mapping_base.get(c) in df_in.columns else ''
                    user_map[c] = st.selectbox(c, choices, index=(choices.index(default) if default in choices else 0), key=f"map_{c}_triune")
                save_map = st.form_submit_button("Apply mapping")
            if save_map:
                st.session_state.triune_mapping = user_map
                st.success("Mapping applied and saved.")

            mapping_use = st.session_state.triune_mapping or mapping_base
            preview_build = pd.DataFrame()
            for c in TRIUNE_COLUMNS:
                src = mapping_use.get(c, "")
                if src and src in df_in.columns:
                    preview_build[c] = df_in[src].astype(str).fillna('')
                else:
                    preview_build[c] = ''

            st.markdown("**Preview of mapped data (first 200 rows)**")
            st.dataframe(preview_build.head(200), height=460, use_container_width=True)

            # Manual Product editor (textarea only)
            render_manual_product_editor(preview_build)

            # Per-product tag editor (textarea only)
            st.markdown("### Per-product TAG ordering")
            st.caption("Edit tag sequence individually per product. Click Apply in the editor to save.")
            render_manual_per_product_tag_editor(preview_build)

            # Generate button
            if st.button("Generate Triune files (preview + download)", key="gen_triune"):
                with st.spinner("Building Triune takeoff..."):
                    manual_products_to_use = st.session_state.get('triune_manual_products')
                    manual_tags_to_use = st.session_state.get('triune_manual_tags_by_product') or {}
                    formatted = takeoff_pipeline(
                        df_in,
                        mapping_use,
                        sort_keys=st.session_state.get('triune_sort_keys', ["PRODUCT","MODEL","TAG"]),
                        ascending=st.session_state.get('triune_sort_ascending', True),
                        manual_products=manual_products_to_use,
                        manual_tags=manual_tags_to_use
                    )
                    st.session_state.formatted_df_triune = formatted
                    st.session_state.ran_triune = True

            if st.session_state.get('ran_triune') and st.session_state.get('formatted_df_triune') is not None:
                formatted = st.session_state.formatted_df_triune
                st.subheader("Final Triune preview (first 500 rows)")
                st.dataframe(formatted.head(500), height=600, use_container_width=True)

                st.markdown("### Download naming")
                if "selected_customer_triune" not in st.session_state:
                    st.session_state.selected_customer_triune = "(None)"
                if "project_name_triune" not in st.session_state:
                    st.session_state.project_name_triune = ""
                with st.form("naming_form_triune", clear_on_submit=False):
                    customer = st.selectbox(
                        "Customer (optional)",
                        options=["(None)"] + list(CUSTOMER_ABBR.keys()),
                        index=0,
                        key="customer_select_triune"
                    )
                    project_name = st.text_input("Project Name (optional)", value=st.session_state.project_name_triune, key="project_input_triune")
                    applied = st.form_submit_button("Apply Name")
                if applied:
                    st.session_state.selected_customer_triune = customer
                    st.session_state.project_name_triune = project_name

                if st.session_state.selected_customer_triune != "(None)" and st.session_state.project_name_triune.strip():
                    base = f"Takeoff_{CUSTOMER_ABBR[st.session_state.selected_customer_triune]}-{st.session_state.project_name_triune.strip()}"
                else:
                    base = uploaded.name.rsplit(".",1)[0] + "_Triune"
                base_safe = sanename(base)

                xbytes = style_to_excel_bytes(formatted,
                                              header_hex=header_color,
                                              model_hex=model_color,
                                              product_hex=product_color,
                                              zebra_hex=zebra_color,
                                              qty_gold_hex=qty_color,
                                              bold_cols=bold_cols)
                csv_bytes = formatted.to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 Download styled Excel", data=xbytes,
                                   file_name=f"{base_safe}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.download_button("📥 Download CSV", data=csv_bytes,
                                   file_name=f"{base_safe}.csv", mime="text/csv")
    else:
        st.info("Upload a CSV or Excel file to begin the Triune conversion.")

# ========== TAB 1: Data Unit Sheet ==========
with tabs[1]:
    st.header("Data Unit Sheet Builder — pick units and export")
    col1, col2 = st.columns(2)
    with col1:
        raw_file = st.file_uploader("📥 Upload Raw Takeoff (CSV or Excel)", type=["csv", "xlsx"], key="unit_raw")
    with col2:
        unit_file = st.file_uploader("📊 Upload Unit Matrix (CSV or Excel)", type=["csv", "xlsx"], key="unit_matrix")

    if not raw_file or not unit_file:
        st.info("Upload both files to continue.")
    else:
        raw_df = load_file_to_df_simple(raw_file)
        unit_df = load_file_to_df_simple(unit_file)
        if raw_df is None or unit_df is None:
            st.error("One of the uploaded files couldn't be read. Please check them.")
        else:
            raw_df = apply_raw_column_mapping(raw_df)
            raw_unit_detected = detect_unit_column(raw_df)
            unit_matrix_clean, matrix_unit_detected = clean_unit_matrix(unit_df)

            st.markdown("### Detected columns")
            st.write(f"Raw unit column (auto-detected): **{raw_unit_detected or 'None'}**")
            st.write(f"Unit matrix unit column (auto-detected): **{matrix_unit_detected}**")

            raw_unit_col = st.selectbox("Select unit column in raw file", options=list(raw_df.columns),
                                        index=list(raw_df.columns).index(raw_unit_detected) if raw_unit_detected in raw_df.columns else 0,
                                        key="raw_unit_col")
            matrix_unit_col = st.selectbox("Select unit column in unit matrix", options=list(unit_matrix_clean.columns),
                                           index=list(unit_matrix_clean.columns).index(matrix_unit_detected) if matrix_unit_detected in unit_matrix_clean.columns else 0,
                                           key="matrix_unit_col")

            guessed_mult = guess_multiplier_column(unit_matrix_clean, matrix_unit_col)
            multiplier_col = st.selectbox("Select multiplier column in unit matrix (units per type)", options=[None] + list(unit_matrix_clean.columns),
                                         index=(1 + list(unit_matrix_clean.columns).index(guessed_mult) if guessed_mult in unit_matrix_clean.columns else 0), key="multcol")
            if multiplier_col is None:
                st.warning("No multiplier column selected — multipliers default to 1 when unmatched.")

            units_series = unit_matrix_clean[matrix_unit_col].dropna().astype(str).str.strip()
            raw_units = raw_df[raw_unit_col].fillna("").astype(str).str.strip().unique().tolist()
            has_empty = any(u == "" for u in raw_df[raw_unit_col].fillna("").astype(str).str.strip().tolist())
            units_list = sorted(set(units_series.tolist() + [u for u in raw_units if u != ""]))
            if has_empty:
                units_list = ["<<EMPTY UNIT>>"] + units_list

            st.markdown("### Choose units to include (checkboxes)")
            select_all = st.checkbox("Select all units", value=True, key="selall")
            cols_chk = st.columns(3)
            selected_units = []
            for i, unit in enumerate(units_list):
                col = cols_chk[i % 3]
                default_val = True if select_all else False
                checked = col.checkbox(str(unit), value=default_val, key=f"unit_chk_{i}")
                if checked:
                    selected_units.append(unit)

            include_empty = st.checkbox("Include empty-unit rows (as <<EMPTY UNIT>>)", value=True, key="inc_empty")
            split_by_unit = st.checkbox("Split into separate Excel per unit (download as ZIP)", value=False, key="splitzip")

            st.markdown("### Output filename")
            default_name = "Data_Unit_Sheet"
            if split_by_unit:
                default_name = "takeoff_by_unit"
            file_name_input = st.text_input("Enter desired download filename (without extension):", value=default_name, key="outfilebase")
            def sanitize_filename_local(s: str) -> str:
                return "".join(c for c in s if c.isalnum() or c in " -_").strip() or default_name
            out_file_base = sanitize_filename_local(file_name_input)

            if st.button("🔁 Generate Data Unit Sheet", key="gen_unit"):
                with st.spinner("Building Data Unit Sheet..."):
                    sel_units = selected_units if selected_units else None
                    final_df = build_data_unit_sheet(
                        raw_df=raw_df,
                        unit_df=unit_matrix_clean,
                        raw_unit_col=raw_unit_col,
                        matrix_unit_col=matrix_unit_col,
                        multiplier_col=multiplier_col,
                        selected_units=sel_units,
                        include_empty=include_empty,
                        default_multiplier=1
                    )
                    st.success("✅ Data Unit Sheet created.")
                    st.write("Preview (first 200 rows):")
                    st.dataframe(final_df.head(200), use_container_width=True)

                    if split_by_unit:
                        groups = {u: g for u, g in final_df.groupby("UNITS")}
                        st.write("Files to be included in ZIP:")
                        summary = [{"unit": k, "rows": len(v)} for k, v in groups.items()]
                        st.table(pd.DataFrame(summary).sort_values("rows", ascending=False))
                        zip_bytes = create_zip_bytes_from_map(groups, export_styled_excel_bytes)
                        download_name = f"{out_file_base}.zip"
                        st.download_button("📦 Download ZIP (per-unit Excels)", data=zip_bytes, file_name=download_name, mime="application/zip")
                    else:
                        excel_bytes = export_styled_excel_bytes(final_df)
                        download_name = f"{out_file_base}.xlsx"
                        st.download_button("💾 Download combined Excel", data=excel_bytes, file_name=download_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
